"""
Exporter
Exportación de listados a archivos planos (CSV y Excel .xlsx)

Recibe filas como diccionarios ordenados (clave = encabezado) y produce:
  - CSV  : texto UTF-8 con BOM, para que Excel interprete los acentos bien
  - XLSX : libro de cálculo real (openpyxl) con encabezado resaltado,
           paneles congelados y anchos de columna automáticos

La extensión del archivo de destino determina el formato elegido.
"""

import csv
import re
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Sequence

from src.utils.helpers import ensure_directory_exists, format_date

# Caracteres de control ilegales en celdas XLSX (todo menos tab/CR/LF)
_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _valor_plano(valor):
    """Normaliza un valor de celda a texto, número o fecha legible"""
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return "Sí" if valor else "No"
    if isinstance(valor, (date, datetime)):
        return format_date(valor)
    if isinstance(valor, float):
        return round(valor, 2)
    if hasattr(valor, "value"):  # Enums
        return valor.value
    if isinstance(valor, str):
        # openpyxl rechaza caracteres de control dentro de las celdas
        return _CONTROL_CHARS.sub(" ", valor)
    return valor


def _encabezados(datos: Sequence[Dict]) -> List[str]:
    """Encabezados de columna en el orden de la primera fila"""
    if not datos:
        raise ValueError("No hay datos para exportar")
    primera = datos[0]
    return [str(clave) for clave in primera.keys()]


def _filas_normalizadas(datos: Sequence[Dict]) -> List[List]:
    """Convierte cada diccionario en una fila alineada con los encabezados"""
    encabezados = _encabezados(datos)
    filas = []
    for registro in datos:
        filas.append([_valor_plano(registro.get(clave, "")) for clave in encabezados])
    return filas


def escribir_csv(datos: Sequence[Dict], ruta: str) -> str:
    """
    Escribe un CSV con codificación UTF-8 (BOM).

    Args:
        datos: Lista de diccionarios (clave = encabezado)
        ruta:  Ruta de salida

    Returns:
        Ruta del archivo generado
    """
    filas = _filas_normalizadas(datos)
    destino = Path(ruta)
    if destino.parent != Path("."):
        ensure_directory_exists(str(destino.parent))

    with open(destino, "w", newline="", encoding="utf-8-sig") as archivo:
        escritor = csv.writer(archivo, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        escritor.writerow(_encabezados(datos))
        escritor.writerows(filas)
    return str(destino)


def escribir_xlsx(datos: Sequence[Dict], ruta: str, hoja: str = "Datos") -> str:
    """
    Escribe un libro Excel (.xlsx) con openpyxl.

    Args:
        datos: Lista de diccionarios (clave = encabezado)
        ruta:  Ruta de salida
        hoja:  Nombre de la hoja de cálculo

    Returns:
        Ruta del archivo generado
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    encabezados = _encabezados(datos)
    filas = _filas_normalizadas(datos)

    destino = Path(ruta)
    if destino.parent != Path("."):
        ensure_directory_exists(str(destino.parent))

    libro = Workbook()
    pagina = libro.active
    pagina.title = hoja[:31] or "Datos"

    pagina.append(encabezados)
    for fila in filas:
        pagina.append(fila)

    # Estilo del encabezado
    for celda in pagina[1]:
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center")

    # Formato numérico para montos y anchos automáticos
    for fila in pagina.iter_rows(min_row=2, max_col=len(encabezados)):
        for celda in fila:
            if isinstance(celda.value, float):
                celda.number_format = "#,##0.00"

    for indice, _ in enumerate(encabezados, start=1):
        letra = get_column_letter(indice)
        ancho = max(
            (len(str(celda.value or "")) for celda in pagina[letra]),
            default=10,
        )
        pagina.column_dimensions[letra].width = min(max(ancho + 2, 10), 60)

    pagina.freeze_panes = "A2"

    libro.save(destino)
    return str(destino)


def exportar_archivo(datos: Sequence[Dict], ruta: str, hoja: str = "Datos") -> str:
    """
    Exporta datos al formato indicado por la extensión de la ruta.

    Args:
        datos: Lista de diccionarios (clave = encabezado)
        ruta:  Ruta de salida (.csv o .xlsx)
        hoja:  Nombre de hoja para archivos Excel

    Returns:
        Ruta del archivo generado
    """
    extension = Path(ruta).suffix.lower()
    if extension == ".csv":
        return escribir_csv(datos, ruta)
    if extension == ".xlsx":
        return escribir_xlsx(datos, ruta, hoja)
    raise ValueError(f"Formato de exportación no soportado: {extension or '(sin extensión)'}")
